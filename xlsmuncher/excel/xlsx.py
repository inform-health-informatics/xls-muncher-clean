"""
Module that provides the concrete implementation for xlsx files
"""

from datetime import date

import openpyxl

from openpyxl.utils import coordinate_to_tuple

from .common import hex2rgb, rgb2hex

class Workbook:
    """Our abstraction of an Excel file (workbook)."""

    def __init__(self, file_path):
        """Construct a Workbook object by opening an Excel file for reading.

        :param file_path: path to the Excel file. The file extension will be used
            to determine format
        """
        self.book = openpyxl.load_workbook(file_path, guess_types=True)

    def sheets(self):
        """A generator over this workbook's sheets."""
        for sheet in self.book:
            yield Sheet(self, sheet)

    def convert_colour(self, colour):
        """Convert from theme+tint colour spec to RGB hex string.

        This method is specific to openpyxl.

        :param colour: an openpyxl.styles.colors.Color instance.
            If colour.rgb is not None, this will be returned directly.
            Otherwise, colour.theme & colour.tint will be used to calculate RGB.
        """
        if colour is None:
            return 'ffffff'
        if colour.type == 'rgb':
            return colour.rgb[2:]  # Strip alpha
        if colour.type == 'indexed':
            from openpyxl.styles.colors import COLOR_INDEX
            return COLOR_INDEX[colour.indexed][2:]  # Strip alpha
        assert colour.type == 'theme'
        new_colour = hex2rgb(self.theme_colours[colour.theme])
        if colour.tint < 0:
            # Darken (shade)
            new_colour = [intensity * (1 + colour.tint) for intensity in new_colour]
        elif colour.tint > 0:
            # Tint (lighten)
            new_colour = [intensity + (255 - intensity) * colour.tint for intensity in new_colour]
        new_colour = tuple(int(round(i)) for i in new_colour)
        return rgb2hex(new_colour)

    @property
    def theme_colours(self):
        """Lazily-loaded theme colour info for a .xlsx file."""
        if not hasattr(self, '_theme_colours'):
            from openpyxl.xml.functions import fromstring, QName
            root = fromstring(self.book.loaded_theme)
            ns = 'http://schemas.openxmlformats.org/drawingml/2006/main'
            theme_elt = root.find(QName(ns, 'themeElements').text)
            colour_schemes = theme_elt.findall(QName(ns, 'clrScheme').text)
            colours = self._theme_colours = []
            for cname in ['lt1', 'dk1', 'lt2', 'dk2',
                          'accent1', 'accent2', 'accent3', 'accent4', 'accent5', 'accent6']:
                c = colour_schemes[0].find(QName(ns, cname).text)
                colour_def = c.getchildren()[0].attrib
                if 'window' in colour_def['val']:
                    colours.append(colour_def['lastClr'])
                else:
                    colours.append(colour_def['val'])
        return self._theme_colours


class Sheet:
    """Our abstraction of a single sheet within a workbook."""

    def __init__(self, workbook, sheet):
        """Create a wrapper around a sheet within a workbook.

        :param workbook: a Workbook instance containing the sheet
        :param sheet: the Excel library's sheet instance
        """
        self.workbook = workbook
        self.book = workbook.book
        self.sheet = sheet

    @property
    def title(self):
        """The sheet's title."""
        return self.sheet.title

    def rows(self, skip_rows=0):
        """A generator over this workbook's rows."""
        for row in self.sheet.iter_rows(min_row=skip_rows + 1):
            yield Row(self, row)

    def row(self, idx):
        """Get a single row from this sheet.

        :param idx: the 0-based index of the row to get
        """
        return Row(self, self.sheet[idx + 1])

    def cell(self, row=None, col=None, *, name=None):
        """Get a single cell from this sheet.

        :param row: the 0-based index of the row to access
        :param col: the 0-based index of the column to access
        """
        if name is not None:
            assert row is None and col is None
            row, col = coordinate_to_tuple(name)
            row -= 1
            col -= 1
        return Cell(self, self.sheet.cell(row=row + 1, column=col + 1))

class Row:
    """Our abstraction of a row in a spreadsheet."""

    def __init__(self, sheet, row):
        """Create a wrapper around a row within a sheet.

        :param sheet: the Sheet instance containing the row
        :param row: the Excel library's row instance
        """
        self.sheet = sheet
        self.book = sheet.book
        self.row = row

    def __len__(self):
        """Get the number of cells in this row."""
        return len(self.row)

    def cells(self):
        """A generator over individual cells in this row."""
        for cell in self.row:
            yield Cell(self.sheet, cell)

    def cell(self, idx):
        """Get a single cell from this row.

        :param idx: the 0-based index of the cell to get
        """
        return Cell(self.sheet, self.row[idx])


class Cell:
    """Our abstraction of a single cell in a spreadsheet."""

    def __init__(self, sheet, cell):
        """Create a wrapper around a cell within a sheet.

        :param sheet: the Sheet instance containing the row
        :param cell: the Excel library's cell instance
        """
        self.sheet = sheet
        self.book = sheet.book
        self.cell = cell

    @property
    def value(self):
        """Get the value of this cell as a Python object.

        Will automatically convert Excel cells marked as string, number, date, or boolean.
        """
        value = self.cell.value
        return value

    @property
    def row(self):
        """The 0-based row index of this cell."""
        return self.cell.row - 1


    @property
    def col(self):
        """The 0-based column index of this cell."""
        return self.cell.col_idx - 1

    @property
    def font(self):
        """Font information for this cell."""
        return Font(self)


class Font:
    """Our abstraction of font information for a cell."""

    def __init__(self, cell):
        """Create a wrapper for font information about a cell.

        :param cell: the cell to provide font information for
        """
        self.book = cell.book
        self.sheet = cell.sheet
        self.cell = cell.cell

    @property
    def strike(self):
        """Whether the cell has strikethrough formatting."""
        return getattr(self.cell.font, 'strike', False)

    @property
    def background(self):
        """The cell's background colour as a string hex code."""
        raise NotImplementedError

    @property
    def foreground(self):
        """The cell's font colour as a string hex code."""
        return self.sheet.workbook.convert_colour(self.cell.font.color)

    @property
    def red(self):
        """Whether the main colour for the cell's text is red."""
        raise NotImplementedError
